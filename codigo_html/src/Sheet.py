from openpyxl import Workbook

class Sheet:

    def __init__(self, sheet_name, file_name, object_sheet, sheet):
        self.sheet_name   = sheet_name
        self.file_name    = file_name
        self.object_sheet = object_sheet
        self.sheet        = sheet

    def set_header1(self):
        self.sheet.title = self.sheet_name
        self.sheet['A1'] = 'URL'
        self.sheet['B1'] = 'Tipo de página'
        self.sheet['C1'] = 'Dados encontrados'
        self.sheet['D1'] = 'Pontuação (K)'
        self.sheet['E1'] = 'Pontuação (N)'
        self.sheet['F1'] = 'Pontuação (K x N)'
        self.sheet['G1'] = 'Pontuação (S)'
        self.sheet['H1'] = 'Pontuação (MP)'
        self.sheet['I1'] = 'Pontuação (Simu)'
        self.sheet['J1'] = 'Validação'        

    def set_header2(self):
        self.sheet.title = self.sheet_name
        self.sheet['A1'] = 'URL'
        self.sheet['B1'] = 'Tipo de página'
        self.sheet['C1'] = 'Dados encontrados'
        self.sheet['D1'] = 'Pontuação (K)'
        self.sheet['E1'] = 'Pontuação (S)'
        self.sheet['F1'] = 'Pontuação (Simu)'
        self.sheet['G1'] = 'Validação'        

    def readable_page1(self, row, url, mime_type, kw_string, sc_list, sc_dic, sc_multiply, sc_severity, sc_wgh_avg, sc_simultaneity):
        self.sheet.cell(row=row, column=1, value=url)
        self.sheet.cell(row=row, column=2, value=mime_type)
        self.sheet.cell(row=row, column=3, value=kw_string)
        self.sheet.cell(row=row, column=4, value=sc_list)
        self.sheet.cell(row=row, column=5, value=sc_dic)
        self.sheet.cell(row=row, column=6, value=sc_multiply)
        self.sheet.cell(row=row, column=7, value=sc_severity)
        self.sheet.cell(row=row, column=8, value=sc_wgh_avg)
        self.sheet.cell(row=row, column=9, value=sc_simultaneity)

    def readable_page2(self, row, url, mime_type, kw_string, sc_list, sc_severity, sc_simultaneity):
        self.sheet.cell(row=row, column=1, value=url)
        self.sheet.cell(row=row, column=2, value=mime_type)
        self.sheet.cell(row=row, column=3, value=kw_string)
        self.sheet.cell(row=row, column=4, value=sc_list)
        self.sheet.cell(row=row, column=5, value=sc_severity)
        self.sheet.cell(row=row, column=6, value=sc_simultaneity) 

    def set_header_others(self, validator):
        if validator == 1:
            self.sheet['K1'] = 'Outros'
        else:
            self.sheet['H1'] = 'Outros'

    def set_content_others(self, row, validator):
        if validator == 1:
            self.sheet.cell(row=row, column=11, value='Form')
        else:
            self.sheet.cell(row=row, column=8, value='Form')

    def not_readable_page(self, row, url, mime_type):
        self.sheet.cell(row=row, column=1, value=url)
        self.sheet.cell(row=row, column=2, value=mime_type)
        self.sheet.cell(row=row, column=3, value='O algoritmo não lê os dados de páginas do tipo ' + mime_type)
    
    def forbidden_page(self, row, url):
        self.sheet.cell(row=row, column=1, value=url)
        self.sheet.cell(row=row, column=2, value='Error')
        self.sheet.cell(row=row, column=3, value='Houve problema na requisição')
        self.sheet.cell(row=row, column=4, value=0)
        self.sheet.cell(row=row, column=5, value=0)
        self.sheet.cell(row=row, column=6, value=0)        

    def save_sheet(self):
        self.object_sheet.save(self.file_name + '.xlsx')